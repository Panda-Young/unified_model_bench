import pandas as pd
import os
import argparse
import sys
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from decimal import Decimal

DEFAULT_HIDE_COLS = [
    "model_input_shape",
    "input_elements",
    "model_output_shape",
    "output_elements",
    "total_run_ms",
    "max_run_idx",
    "init_ms",
    "app_name",
    "notes",
]

MERGE_COLS = [
    "date",
    "time",
    "model_name",
    "model_input_shape",
    "input_elements",
    "model_output_shape",
    "output_elements",
    "device_info",
    "arch",
    "app_name",
]


def format_device_info(text):
    """Format device_info by splitting CPU and GPU information on separate lines."""
    if not text or pd.isna(text):
        return ""
    text = str(text).strip()
    if not text:
        return ""
    if "\n" in text:
        return text

    parts = []
    for delimiter in [";", ",", "|"]:
        if delimiter in text:
            parts = [p.strip() for p in text.split(delimiter) if p.strip()]
            break
    if not parts:
        parts = [text]

    cpu_parts, gpu_parts = [], []
    gpu_keywords = ["gpu", "adreno", "mali", "nvidia", "amd", "powervr", "vivante", "graphics"]
    for part in parts:
        if any(kw in part.lower() for kw in gpu_keywords):
            gpu_parts.append(part)
        else:
            cpu_parts.append(part)

    result = []
    if cpu_parts:
        result.append("; ".join(cpu_parts))
    result.extend(gpu_parts)
    return "\n".join(result)


def convert_scientific_notation(value):
    if value is None or pd.isna(value):
        return ""
    value_str = str(value).strip()
    if not value_str:
        return ""
    try:
        num = float(value_str)
        if "e" in value_str.lower():
            return format(Decimal(value_str), "f")
        return value_str
    except (ValueError, TypeError):
        return value_str


def csv_to_xlsx(in_csv, out_xlsx=None, hide_cols=None, sheet_name="Sheet1", encoding="utf-8-sig"):
    if out_xlsx is None:
        out_xlsx = os.path.splitext(in_csv)[0] + ".xlsx"

    hide_cols = hide_cols or DEFAULT_HIDE_COLS
    hide_cols_lower = [c.strip().lower() for c in hide_cols]

    # Try multiple encodings: UTF-8 first, then GBK (Chinese Windows locale),
    # then latin-1 as a last resort (never fails, but may mangle characters).
    encodings_to_try = [encoding]
    if encoding != "gbk":
        encodings_to_try.append("gbk")
    if encoding != "gb2312":
        encodings_to_try.append("gb2312")
    encodings_to_try.append("latin-1")

    df = None
    for enc in encodings_to_try:
        try:
            df = pd.read_csv(in_csv, encoding=enc, dtype=str, keep_default_na=False)
            if enc != encoding:
                print(f"Note: CSV read with encoding '{enc}' (fallback from '{encoding}')")
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:
            print(f"Error reading CSV '{in_csv}' with encoding '{enc}': {e}")
            continue

    if df is None:
        print(f"Error: Could not read CSV '{in_csv}' with any encoding.")
        return

    if "device_info" in df.columns:
        df["device_info"] = df["device_info"].apply(format_device_info)

    # Grouping boundaries
    df_group = df.copy()
    for col in ("time", "model_name"):
        if col in df_group.columns:
            df_group[col] = df_group[col].astype(str).fillna("").str.strip()
            df_group[col + "_grp"] = df_group[col].replace("", pd.NA).ffill().bfill().astype(str)

    # Time batch end rows
    time_batch_ends = []
    if "time_grp" in df_group.columns:
        mask = df_group["time_grp"] != df_group["time_grp"].shift(-1)
        ends_idx = df_group.index[mask.fillna(True)].tolist()
        time_batch_ends = [i + 2 for i in ends_idx]

    # Model batch ranges
    model_batch_ends = []
    if "model_name_grp" in df_group.columns:
        ends_mask = (df_group["model_name_grp"] != df_group["model_name_grp"].shift(-1)).fillna(True)
        ends_idx = df_group.index[ends_mask].tolist()
        model_batch_ends = [i + 2 for i in ends_idx]

    # Write to Excel
    try:
        df.to_excel(out_xlsx, index=False, sheet_name=sheet_name, engine="openpyxl")
    except PermissionError:
        print(f"Error: File '{out_xlsx}' is open. Attempting to close Excel...")
        os.system("taskkill /f /im EXCEL.exe")
        time.sleep(0.5)
        try:
            df.to_excel(out_xlsx, index=False, sheet_name=sheet_name, engine="openpyxl")
        except Exception as e:
            print(f"Error writing XLSX: {e}")
            return
    except Exception as e:
        print(f"Error writing XLSX: {e}")
        return

    wb = load_workbook(out_xlsx)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    max_row, max_col = ws.max_row, ws.max_column

    # Convert scientific notation
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                converted = convert_scientific_notation(cell.value)
                if converted != str(cell.value):
                    cell.value = converted
            cell.number_format = "@"

    # Header mapping
    header_to_idx = {}
    for idx in range(1, max_col + 1):
        h = ws.cell(row=1, column=idx).value
        if h:
            header_to_idx[str(h).strip().lower()] = idx

    merge_col_indices = [header_to_idx.get(c.lower()) for c in MERGE_COLS]
    merge_col_indices = [i for i in merge_col_indices if i is not None]
    device_info_idx = header_to_idx.get("device_info")

    # Merge adjacent identical cells
    for col_idx in merge_col_indices:
        row = 2
        while row <= max_row:
            start_row = row
            start_value = str(ws.cell(row=start_row, column=col_idx).value or "").strip()
            while row <= max_row:
                cur = str(ws.cell(row=row, column=col_idx).value or "").strip()
                if cur != start_value:
                    break
                row += 1
            end_row = row - 1
            if end_row > start_row:
                try:
                    ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                    cell = ws.cell(row=start_row, column=col_idx)
                    if col_idx == device_info_idx:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                except Exception:
                    pass

    # Borders at batch boundaries
    medium_border = Border(bottom=Side(style="medium", color="000000"))
    thick_border = Border(bottom=Side(style="thick", color="000000"))
    for row_num in model_batch_ends:
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_num, column=col_idx).border = thick_border
    for row_num in time_batch_ends:
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_num, column=col_idx).border = medium_border

    # Column widths and alignment
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if col_idx == device_info_idx:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif not any(cell.coordinate in str(m) for m in ws.merged_cells):
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            if cell.value:
                text = str(cell.value)
                if col_idx == device_info_idx and "\n" in text:
                    max_len = max(max_len, max(len(line) for line in text.split("\n")))
                else:
                    max_len = max(max_len, len(text))
        ws.column_dimensions[col_letter].width = max(8, min(max_len + 2, 120))

    # Hide columns
    for idx in range(1, max_col + 1):
        h = ws.cell(row=1, column=idx).value
        if h and str(h).strip().lower() in hide_cols_lower:
            ws.column_dimensions[get_column_letter(idx)].hidden = True

    ws.freeze_panes = "A2"

    try:
        wb.save(out_xlsx)
        print(f"Saved: {out_xlsx}")
    except PermissionError:
        print(f"Error: Unable to save '{out_xlsx}'. Please close the file and try again.")
        return

    if sys.platform == "win32":
        try:
            os.startfile(out_xlsx)
        except Exception:
            pass


def parse_args():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_csv = os.path.normpath(os.path.join(script_dir, "..", "summary.csv"))
    parser = argparse.ArgumentParser(description="Convert CSV to formatted XLSX")
    parser.add_argument("csv", nargs="?", default=default_csv, help="Input CSV file")
    parser.add_argument("-o", "--out", help="Output XLSX file")
    parser.add_argument("-s", "--sheet", default="Sheet1", help="Sheet name")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    csv_to_xlsx(in_csv=args.csv, out_xlsx=args.out, sheet_name=args.sheet)
